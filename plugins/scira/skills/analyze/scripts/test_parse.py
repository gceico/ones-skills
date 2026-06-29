#!/usr/bin/env python3
"""Self-check for parse_xlsx: builds a demo LinkedIn-shaped xlsx, parses it twice, asserts idempotence.

Run: python3 test_parse.py   (needs openpyxl)
"""
import csv, io, json, tempfile, contextlib
from pathlib import Path
import openpyxl
import parse_xlsx


def make_demo(path):
    wb = openpyxl.Workbook()
    tp = wb.active
    tp.title = "TOP POSTS"
    tp.append(["Top posts", None, None, None, None, None, None])      # row 1 note
    tp.append([None] * 7)                                             # row 2
    tp.append(["Post URL", "Date", "Engagements", None, "Post URL", "Date", "Impressions"])  # row 3 headers
    tp.append(["https://www.linkedin.com/feed/update/1", "2026-05-02", 120, None,
               "https://www.linkedin.com/feed/update/1", "2026-05-02", 4000])
    tp.append(["https://www.linkedin.com/feed/update/2", "2026-05-05", 30, None,
               "https://www.linkedin.com/feed/update/2", "2026-05-05", 2000])
    fol = wb.create_sheet("FOLLOWERS")
    fol.append(["Total followers", 1500])
    fol.append([None, None])
    fol.append(["2026-05-01", 12])
    fol.append(["2026-05-02", 8])
    demo = wb.create_sheet("DEMOGRAPHICS")
    demo.append(["Job title", "Percentage"])
    demo.append(["Software Engineer", 0.31])
    demo.append(["Founder", 0.18])
    wb.save(path)


def run(xlsx, data_dir, period):
    buf = io.StringIO()
    with contextlib.redirect_stdout(buf):
        parse_xlsx.main([str(xlsx), "--data-dir", str(data_dir), "--period", period])
    return json.loads(buf.getvalue())


def rowcount(data_dir):
    f = Path(data_dir) / "analytics.csv"
    with f.open(newline="") as fh:
        return sum(1 for _ in csv.DictReader(fh))


def main():
    with tempfile.TemporaryDirectory() as d:
        d = Path(d)
        xlsx = d / "Content_2026-05-01_2026-05-31_Demo.xlsx"
        make_demo(xlsx)

        r1 = run(xlsx, d, "2026-05")
        assert r1["posts_parsed"] == 2, r1
        assert r1["rows_appended"] == 2, r1
        assert rowcount(d) == 2, "first import should write 2 rows"

        # engagement rate: 120/4000*100 = 3.0
        ers = {p["post_url"]: p["engagement_rate"] for p in r1["posts"]}
        assert ers["https://www.linkedin.com/feed/update/1"] == 3.0, ers

        r2 = run(xlsx, d, "2026-05")          # re-import same file+period
        assert r2["rows_appended"] == 0, r2
        assert r2["rows_skipped_duplicate"] == 2, r2
        assert rowcount(d) == 2, "re-import must NOT duplicate rows"

        assert r1["followers"]["total"] == 1500, r1["followers"]
        print("OK — parse + idempotent de-dupe verified")


if __name__ == "__main__":
    main()
