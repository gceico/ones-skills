#!/usr/bin/env python3
"""Parse a LinkedIn analytics xlsx, append per-post rows to analytics.csv, print a JSON summary.

Idempotent: re-running with the same file + period does not duplicate rows (de-dupe key = post_url+period).
Sheet layout is documented in ../../../references/conventions.md.
"""
import argparse, csv, json, os, sys, re
from pathlib import Path

HEADER = ["period", "post_url", "publish_date", "impressions", "engagements",
          "engagement_rate", "topic", "format", "hook_type", "tier"]


def _cell(v):
    return "" if v is None else str(v).strip()


def _num(v):
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, AttributeError):
        return None


def _looks_url(v):
    return isinstance(v, str) and "linkedin.com" in v


def parse_top_posts(ws):
    """Merge the side-by-side engagements (cols A-C) and impressions (cols E-G) tables by Post URL."""
    eng, imp = {}, {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        # left table: url, date, engagements (cols 0-2)
        if len(row) >= 3 and _looks_url(row[0]):
            eng[row[0].strip()] = (_cell(row[1]), _num(row[2]))
        # right table: url, date, impressions (cols 4-6)
        if len(row) >= 7 and _looks_url(row[4]):
            imp[row[4].strip()] = _num(row[6])
    posts = []
    for url, (date, engagements) in eng.items():
        impressions = imp.get(url)
        er = round(engagements / impressions * 100, 1) if engagements and impressions else None
        posts.append({"post_url": url, "publish_date": date,
                      "impressions": impressions, "engagements": engagements,
                      "engagement_rate": er})
    return posts


def parse_followers(wb):
    if "FOLLOWERS" not in wb.sheetnames:
        return {"total": None, "new_series": []}
    ws = wb["FOLLOWERS"]
    rows = list(ws.iter_rows(values_only=True))
    total = _num(rows[0][1]) if rows and len(rows[0]) > 1 else (_num(rows[0][0]) if rows else None)
    series = [(_cell(r[0]), _num(r[1])) for r in rows[2:] if r and len(r) > 1 and _num(r[1]) is not None]
    return {"total": total, "new_series": series}


def parse_demographics(wb):
    if "DEMOGRAPHICS" not in wb.sheetnames:
        return []
    out = []
    for r in wb["DEMOGRAPHICS"].iter_rows(values_only=True):
        if r and len(r) >= 2 and _cell(r[0]) and _num(r[1]) is not None:
            out.append((_cell(r[0]), _num(r[1])))
    return out


def load_swipe_enrichment(data_dir):
    """post_url -> (topic, format, hook_type) from swipe-file.csv if present."""
    f = Path(data_dir) / "swipe-file.csv"
    if not f.exists():
        return {}
    out = {}
    with f.open(newline="", encoding="utf-8") as fh:
        for row in csv.DictReader(fh):
            url = (row.get("post_url") or "").strip()
            if url:
                out[url] = ((row.get("theme_tags") or "").split("|")[0],
                            row.get("media_type") or "", row.get("hook_pattern") or "")
    return out


def existing_keys(csv_path):
    keys = set()
    if csv_path.exists():
        with csv_path.open(newline="", encoding="utf-8") as fh:
            for row in csv.DictReader(fh):
                keys.add((row.get("post_url", ""), row.get("period", "")))
    return keys


def period_from_filename(path):
    m = re.search(r"(\d{4})[-_](\d{2})[-_]\d{2}", os.path.basename(path))
    return f"{m.group(1)}-{m.group(2)}" if m else "unknown"


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--data-dir", required=True)
    ap.add_argument("--period")
    args = ap.parse_args(argv)

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl missing — run: python3 -m pip install --user openpyxl")

    data_dir = Path(os.path.expanduser(args.data_dir))
    data_dir.mkdir(parents=True, exist_ok=True)
    csv_path = data_dir / "analytics.csv"
    period = args.period or period_from_filename(args.xlsx)

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if "TOP POSTS" not in wb.sheetnames:
        sys.exit("No 'TOP POSTS' sheet — not a LinkedIn content export?")
    posts = parse_top_posts(wb["TOP POSTS"])
    enrich = load_swipe_enrichment(data_dir)
    seen = existing_keys(csv_path)

    new_rows = []
    for p in posts:
        key = (p["post_url"], period)
        if key in seen:
            continue
        topic, fmt, hook = enrich.get(p["post_url"], ("", "", ""))
        new_rows.append({"period": period, "post_url": p["post_url"],
                         "publish_date": p["publish_date"],
                         "impressions": p["impressions"] or "",
                         "engagements": p["engagements"] or "",
                         "engagement_rate": p["engagement_rate"] if p["engagement_rate"] is not None else "",
                         "topic": topic, "format": fmt, "hook_type": hook, "tier": ""})
        seen.add(key)

    write_header = not csv_path.exists()
    with csv_path.open("a", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=HEADER)
        if write_header:
            w.writeheader()
        for r in new_rows:
            w.writerow(r)

    print(json.dumps({
        "period": period,
        "posts_parsed": len(posts),
        "rows_appended": len(new_rows),
        "rows_skipped_duplicate": len(posts) - len(new_rows),
        "followers": parse_followers(wb),
        "demographics_top": parse_demographics(wb)[:10],
        "analytics_csv": str(csv_path),
        "posts": new_rows,
    }, indent=2, default=str))


if __name__ == "__main__":
    main()
